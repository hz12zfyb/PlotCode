# -*- coding: utf-8 -*-
"""
Created on Wed Mar  4 23:20:47 2020
根据当前的保留时间预测接下一串保留时间
@author: Administrator
"""

import numpy as np
import openpyxl
import sys
sys.path.append('D:/Refresh/py36')
import support_pie
import win32clipboard as w
import win32con
 
def SetCutBoardStr(aString):
    w.OpenClipboard()
    w.EmptyClipboard()
    w.SetClipboardData(win32con.CF_UNICODETEXT, aString)
    w.CloseClipboard()
 
excel_dir = 'D:/Refresh/data/南极色素-20201028-冯毓彬/text.xlsx'
file_dir = 'D:/Refresh/data/南极色素-20201028-冯毓彬/pdfplumber_txt_440/'

#station_index = int(input("输入旧样品序号:"))
station_index += 1

#读取excel
wb = openpyxl.load_workbook(excel_dir)
sheet_r = wb['station']


station_y = 1 + station_index * 5
station_name_old = sheet_r.cell(1,station_y).value
station_name_new = sheet_r.cell(1,station_y+5).value     
txtnames = support_pie.file_name(file_dir,'.txt')
data = np.loadtxt(txtnames[station_index],dtype='float',delimiter=',')[:,1]
station_times_old,station_times_new = np.zeros(27),np.full(27, np.nan)
for i in range(len(station_times_old)):
    station_times_old[i] = sheet_r.cell(i+6,station_y-1).value
    if station_times_old[i] != np.nan:
        res = np.abs(data-station_times_old[i])
        if np.min(res) <= 0.1:
            station_times_new[i] = data[np.argmin(res)]

#复制到剪贴板
text =''
for i in range(27):
    if station_times_new[i] > 0 :
        text += str(station_times_new[i])+'\n'
    else:
        text += '\n'
SetCutBoardStr(text)
print('参照样品名为%s,预测样品名为%s'%(station_name_old,station_name_new))