# -*- coding: utf-8 -*-
"""
Created on Sun Mar 24 18:31:37 2019
从txt里读取数据 写入excel
@author: Administrator
"""
import numpy as np
import openpyxl
import os,sys
sys.path.append('D:/Refresh/py36')
import support_pie

def check_data(data):
    for i in range(data.shape[0]-1):
        if data[i+1,1]-data[i,1] < 0 :
            print('数据有误')
            break
        else:
            continue
    

#读取excel 
excel_dir = 'D:/Refresh/data/南极色素-20201028-冯毓彬/text.xlsx'
file_dir = 'D:/Refresh/data/南极色素-20201028-冯毓彬/pdfplumber_txt_440/'
wb = openpyxl.load_workbook(excel_dir)
#sheet_name = wb.sheetnames
sheet_r = wb['station']
#station_y = 6
station_num = int((sheet_r.max_column-3)/5)
error_names =[]
for i in range(station_num):
    station_y = 6 + i * 5
    station_name = sheet_r.cell(1,station_y).value    
    txtnames = support_pie.file_name(file_dir,'.txt')
    for txtname in txtnames:
        split_ser = txtname.split('-')
        stationname = split_ser[3]
        if station_name == stationname:
#            print(station_name)   
            data = np.loadtxt(txtname,dtype='float',delimiter=',')
            check_data(data)
    
            pig_index =np.linspace(6,32,num=27,dtype='int')
            for j in pig_index: 
                blank  = sheet_r.cell(1,1).value
                res_time = sheet_r.cell(j,station_y-1).value
                pig_area = sheet_r.cell(j,station_y).value
                if res_time != blank and pig_area == blank:
                    res = data[:,1] - res_time
                    if len(np.where(res == 0)[0]) == 1:
                        value = float(data[np.where(res == 0),2][0])
                        sheet_r.cell(j,station_y,value)
                        value = -1
                    else:
#                        wb.save(excel_dir)
                        print('%s保留时间错误：错误值为：%.3f'%(station_name,res_time))

                else:
                    continue           
#        elif txtname == txtnames[-1]:
#            error_names.append(station_name)
#            print('未找到' +station_name +'站位')

  
wb.save(excel_dir) 