# -*- coding: utf-8 -*-
"""
Created on Sun Apr 21 21:57:11 2019
1. 站位填满
2. 读取色素时，填充text.xlsx-station工作表的编号，定容信息
填写站位 层次
3. 乱序查找
@author: Administrator
"""

import openpyxl
import pandas as pd
import numpy as np
import os,sys
sys.path.append('D:/Refresh/py36')
import support_pie


#1
wb = openpyxl.load_workbook('D:/Refresh/data/CHINARE-36/odv/data_all.xlsx')
ws = wb['Sheet1']
blank = ws.cell(1,8).value

for i in range(ws.max_row):
    if ws.cell(i+1,1).value != blank:
        value = ws.cell(i+1,1).value
#        value2 = ws.cell(i+1,3).value
    else :
        ws.cell(i+1,1).value = value
#        ws.cell(i+1,3).value = value2
wb.save('a.xlsx')        

#2
informations =  support_pie.file_name(file_dir='D:/Refresh/data/南极色素-20201028-冯毓彬/440',
                                      file_postfix='.pdf')
df = pd.read_excel('D:/Refresh/data/南极色素-20201028-冯毓彬/采样站位-标号对应.xlsx').\
        set_index('编号')
excel_dir = 'D:/Refresh/data/南极色素-20201028-冯毓彬/text.xlsx'    
wb = openpyxl.load_workbook(excel_dir)
ws = wb['station']
for i in range(int((ws.max_column-3)/5)):
    a = informations[i].split('-')
    j = 5 * i 
    ws.cell(1,j+6).value = a[3] #34th157
    ws.cell(3,j+6).value = int(a[5])/1000000 #200/10^6
    ws.cell(4,j+6).value = int(a[6].replace('ul',''))/1000000 #10/10^6
    ws.cell(1,j+7).value = int(a[2].split('\\')[1]) #2019051510
    ws.cell(3,j+7).value = df.loc[a[3]]['样品名称']
    ws.cell(2,j+6).value = df.loc[a[3]]['过滤体积/L']
wb.save(excel_dir.replace('text','text-1'))

#3
#informations =  support_pie.file_name(file_dir='D:/Refresh/data/南极色素-20190510-冯毓彬/34th阿蒙森海',
#                                      file_postfix='.pdf')
#a = np.zeros(283)
#for i in range(283):
#    a[i] = informations[i].split('-')[3].split('th')[1]
#np.where(a == 18)







